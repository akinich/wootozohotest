import streamlit as st
import pandas as pd
import requests
from datetime import datetime
from dateutil.parser import parse
from collections import Counter
from io import BytesIO
from zipfile import ZipFile
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

# ------------------------
# NEW: Display item database and build mapping
try:
    # Read all columns as strings so HSN leading zeros are preserved
    item_db_df = pd.read_excel("item_database.xlsx", dtype=str)
    # Normalize column headers to lowercase and stripped
    item_db_df.columns = [str(col).strip().lower() for col in item_db_df.columns]

    # Validate required columns (case-insensitive because we normalized)
    required_columns = ["woocommerce name", "zoho name", "hsn", "usage unit"]
    for col in required_columns:
        if col not in item_db_df.columns:
            st.error(f"Missing required column in item_database.xlsx: '{col}'")
            st.stop()

    st.subheader("Item Database")
    st.dataframe(item_db_df)

    # Build mapping: lower(woocommerce name) -> { zoho, hsn, usage_unit }
    name_mapping = {}
    for _, row in item_db_df.iterrows():
        woo_raw = row.get("woocommerce name", "")
        woo = str(woo_raw).strip().lower()
        if not woo:
            continue
        # Only take first match (do not overwrite if key already exists)
        if woo in name_mapping:
            continue
        zoho = "" if pd.isna(row.get("zoho name")) else str(row.get("zoho name")).strip()
        hsn_val = "" if pd.isna(row.get("hsn")) else str(row.get("hsn")).strip()
        usage_val = "" if pd.isna(row.get("usage unit")) else str(row.get("usage unit")).strip()
        name_mapping[woo] = {"zoho": zoho, "hsn": hsn_val, "usage_unit": usage_val}

except FileNotFoundError:
    st.warning("item_database.xlsx not found. Please upload it to the app folder.")
    name_mapping = {}
except Exception as e:
    st.error(f"Error reading item_database.xlsx: {e}")
    name_mapping = {}

# ------------------------
# WooCommerce API settings
WC_API_URL = st.secrets.get("WC_API_URL")
WC_CONSUMER_KEY = st.secrets.get("WC_CONSUMER_KEY")
WC_CONSUMER_SECRET = st.secrets.get("WC_CONSUMER_SECRET")

if not WC_API_URL or not WC_CONSUMER_KEY or not WC_CONSUMER_SECRET:
    missing = [k for k, v in {
        "WC_API_URL": WC_API_URL,
        "WC_CONSUMER_KEY": WC_CONSUMER_KEY,
        "WC_CONSUMER_SECRET": WC_CONSUMER_SECRET
    }.items() if not v]
    st.error("WooCommerce API credentials missing: " + ", ".join(missing))
    st.stop()

st.title("WooCommerce → Accounting CSV & Excel Export Tool")

# ------------------------
# Date input fields
start_date = st.date_input("Start Date")
end_date = st.date_input("End Date")

# Invoice number customization
invoice_prefix = st.text_input("Invoice Prefix", value="ECHE/2526/")
start_sequence = st.number_input("Starting Sequence Number", min_value=1, value=608)

if start_date > end_date:
    st.error("Start date cannot be after end date.")

fetch_button = st.button("Fetch Orders", disabled=(start_date > end_date))

# ------------------------
def to_float(x):
    try:
        if x is None or x == "":
            return 0.0
        return float(x)
    except Exception:
        return 0.0

# ------------------------
if fetch_button:
    st.info("Fetching orders from WooCommerce...")

    start_iso = start_date.strftime("%Y-%m-%dT00:00:00")
    end_iso = end_date.strftime("%Y-%m-%dT23:59:59")

    # Fetch orders with pagination
    all_orders = []
    page = 1
    try:
        with st.spinner("Fetching orders from WooCommerce..."):
            while True:
                response = requests.get(
                    f"{WC_API_URL}/orders",
                    params={
                        "after": start_iso,
                        "before": end_iso,
                        "per_page": 100,
                        "page": page
                    },
                    auth=(WC_CONSUMER_KEY, WC_CONSUMER_SECRET),
                    timeout=30
                )
                response.raise_for_status()
                orders = response.json()
                if not orders:
                    break
                all_orders.extend(orders)
                page += 1
    except requests.exceptions.RequestException as e:
        st.error(f"Error fetching orders: {e}")
        st.stop()

    if not all_orders:
        st.warning("No orders found in this date range.")
        st.stop()

    all_orders.sort(key=lambda x: x["id"])
    status_counts = Counter(order["status"].lower() for order in all_orders)
    def get_status_count(variants): return sum(status_counts.get(v,0) for v in variants)

    # ------------------------
    # Transform completed orders into line-item CSV
    csv_rows = []
    replacements_log = []  # will record replacements made
    sequence_number = start_sequence
    completed_orders = [o for o in all_orders if o["status"].lower() == "completed"]

    # CHANGE: Stop if no completed orders
    if not completed_orders:
        st.warning("No completed orders found in this date range.")
        st.stop()

    for order in completed_orders:
        order_id = order["id"]
        invoice_number = f"{invoice_prefix}{sequence_number:05d}"
        sequence_number += 1
        invoice_date = parse(order["date_created"]).strftime("%Y-%m-%d %H:%M:%S")
        customer_name = f"{order['billing'].get('first_name','')} {order['billing'].get('last_name','')}".strip()
        place_of_supply = order['billing'].get('state', '')
        currency = order.get('currency','')
        shipping_charge = to_float(order.get('shipping_total',0))
        entity_discount = to_float(order.get('discount_total',0))

        for item in order.get("line_items", []):
            product_meta = item.get("meta_data",[]) or []
            # Default - try to pick HSN & usage unit from item's meta_data first
            hsn = ""
            usage_unit = ""
            for meta in product_meta:
                key = str(meta.get("key","")).lower()
                if key == "hsn":
                    hsn_val = meta.get("value","")
                    # Keep as string (preserve whatever format present); do not strip leading zeros
                    hsn = "" if hsn_val is None else str(hsn_val)
                if key == "usage unit":
                    usage_val = meta.get("value","")
                    usage_unit = "" if usage_val is None else str(usage_val)

            # NEW: Replace item name using item_database mapping (case-insensitive)
            original_item_name = item.get("name","")
            item_name_lower = str(original_item_name).strip().lower()
            if item_name_lower in name_mapping:
                mapping = name_mapping[item_name_lower]
                item_name_final = mapping.get("zoho", original_item_name)
                # If replaced, override HSN and Usage unit from item_database.xlsx
                hsn_from_db = mapping.get("hsn", "")
                usage_from_db = mapping.get("usage_unit", "")
                # Preserve as strings (hsn_from_db already read as str because dtype=str)
                if hsn_from_db is not None and hsn_from_db != "":
                    hsn = hsn_from_db
                if usage_from_db is not None and usage_from_db != "":
                    usage_unit = usage_from_db

                # Log replacement
                replacements_log.append({
                    "Original WooCommerce Name": original_item_name,
                    "Replaced Zoho Name": item_name_final,
                    "HSN": hsn,
                    "Usage unit": usage_unit
                })
            else:
                item_name_final = original_item_name

            # CHANGE: Ensure Item Tax % is always numeric
            tax_class = item.get("tax_class") or ""
            try:
                item_tax_pct = float(tax_class)
            except (TypeError, ValueError):
                item_tax_pct = 0.0

            row = {
                "Invoice Number": invoice_number,
                "PurchaseOrder": order_id,
                "Invoice Date": invoice_date,
                "Invoice Status": order["status"].capitalize(),
                "Customer Name": customer_name,
                "Place of Supply": place_of_supply,
                "Currency Code": currency,
                "Item Name": item_name_final,
                "HSN/SAC": hsn,
                "Item Type": item.get("type","goods"),
                "Quantity": item.get("quantity",0),
                "Usage unit": usage_unit,
                "Item Price": to_float(item.get("price",0)),  # CHANGE: Ensure numeric
                "Is Inclusive Tax":"FALSE",
                "Item Tax %": item_tax_pct,  # numeric tax
                "Discount Type":"entity_level",
                "Is Discount Before Tax":"TRUE",
                "Entity Discount Amount":entity_discount,
                "Shipping Charge":shipping_charge,
                "Item Tax Exemption Reason":"ITEM EXEMPT FROM GST",
                "Supply Type":"Exempted",
                "GST Treatment":"consumer"
            }
            csv_rows.append(row)

    df = pd.DataFrame(csv_rows)
    st.dataframe(df.head(50))

    # ------------------------
    # Show Replacements Log
    if replacements_log:
        st.subheader("Item Name Replacements Log")
        st.dataframe(pd.DataFrame(replacements_log))

    # ------------------------
    # Revenue only from WooCommerce totals
    total_revenue_by_order_total = 0.0
    for order in completed_orders:
        order_total = to_float(order.get("total",0))
        refunds = order.get("refunds") or []
        refund_total = sum(to_float(r.get("amount") or r.get("total") or r.get("refund_total") or 0) for r in refunds)
        net_total = order_total - refund_total
        total_revenue_by_order_total += net_total

    first_order_id = completed_orders[0]["id"] if completed_orders else None
    last_order_id = completed_orders[-1]["id"] if completed_orders else None
    first_invoice_number = f"{invoice_prefix}{start_sequence:05d}"
    last_invoice_number = f"{invoice_prefix}{sequence_number-1:05d}" if completed_orders else None

    # ------------------------
    # Summary metrics
    summary_metrics = {
        "Metric":[
            "Total Orders Fetched",
            "Completed Orders",
            "Processing Orders",
            "On Hold Orders",
            "Cancelled Orders",
            "Pending Payment Orders",
            "Completed Order ID Range",
            "Invoice Number Range",
            "Total Revenue (Net of Refunds)"
        ],
        "Value":[
            len(all_orders),
            get_status_count(['completed']),
            get_status_count(['processing']),
            get_status_count(['on-hold','on_hold','on hold']),
            get_status_count(['cancelled','canceled']),
            get_status_count(['pending','pending payment','pending-payment']),
            f"{first_order_id} → {last_order_id}" if completed_orders else "",
            f"{first_invoice_number} → {last_invoice_number}" if completed_orders else "",
            total_revenue_by_order_total
        ]
    }
    summary_df = pd.DataFrame(summary_metrics)

    # ------------------------
    # Order Details sheet
    order_details_rows = []
    sequence_number_temp = start_sequence
    for order in completed_orders:
        invoice_number_temp = f"{invoice_prefix}{sequence_number_temp:05d}"
        sequence_number_temp += 1
        order_total = to_float(order.get("total",0))
        refunds = order.get("refunds") or []
        refund_total = sum(to_float(r.get("amount") or r.get("total") or r.get("refund_total") or 0) for r in refunds)
        net_total = order_total - refund_total
        order_details_rows.append({
            "Invoice Number": invoice_number_temp,
            "Order Number": order["id"],
            "Date": parse(order["date_created"]).strftime("%Y-%m-%d %H:%M:%S"),
            "Customer Name": f"{order['billing'].get('first_name','')} {order['billing'].get('last_name','')}".strip(),
            "Order Total": net_total
        })
    order_details_df = pd.DataFrame(order_details_rows)
    grand_total = order_details_df["Order Total"].sum()
    grand_total_row = {
        "Invoice Number": "Grand Total",
        "Order Number": "",
        "Date": "",
        "Customer Name": "",
        "Order Total": grand_total
    }
    order_details_df = pd.concat([order_details_df, pd.DataFrame([grand_total_row])], ignore_index=True)

    # ------------------------
    # Prepare Excel
    excel_output = BytesIO()
    with pd.ExcelWriter(excel_output, engine='openpyxl') as writer:
        summary_df.to_excel(writer, index=False, sheet_name="Summary Metrics")
        order_details_df.to_excel(writer, index=False, sheet_name="Order Details")
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                max_length = max(len(str(c.value)) if c.value is not None else 0 for c in col) + 2
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_length
    excel_data = excel_output.getvalue()

    # ------------------------
    # CSV
    csv_bytes = df.to_csv(index=False).encode('utf-8')

    # ------------------------
    # Create combined ZIP
    zip_buffer = BytesIO()
    with ZipFile(zip_buffer, "w") as zip_file:
        zip_file.writestr(f"orders_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv", csv_bytes)
        zip_file.writestr(f"summary_report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx", excel_data)

    zip_buffer.seek(0)

    # ------------------------
    # Download ZIP button
    st.download_button(
        label="Download CSV + Excel (Combined ZIP)",
        data=zip_buffer,
        file_name=f"woocommerce_export_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.zip",
        mime="application/zip"
    )
